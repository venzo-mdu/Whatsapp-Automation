from flask import Blueprint
from flask import jsonify,request
import openpyxl
from datetime import date,datetime,timedelta
from flask_apscheduler import APScheduler
import time
from pywhatkit.whats import sendwhats_image,sendwhatmsg_instantly
from . import scheduler
import xlwt
from xlwt import Workbook
from urllib.request import urlopen,Request
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials



views = Blueprint('views',__name__)

# task to perform
def task(*args):
    # xml file to python dictionary
    file_path = args[2]
    image_folder = args[0]
    group_id = args[1]
    holiday_file_path = args[3]
    today = date.today()

    xml_book = openpyxl.load_workbook(file_path)
    xml_file = (xml_book).active
    first_row = list(xml_file.rows)[0]
    keys =[]
    for key in first_row:
        key = (key.value).lower()
        if 'birth' in key or 'dob' in key:
            keys.append('dob')
        else:
            keys.append(key)
    print(keys)
    persons = []
    row_position = 1
    for row in xml_file.iter_rows(2, xml_file.max_row):
        row_position += 1
        xml_data = {}
        for i in range(0,len(row)):
            xml_data[keys[i]] = row[i].value
        xml_data['row'] = row_position
        persons.append(xml_data)

    max_column = xml_file.max_column

    # try:
    #     scheduler.remove_job('check')
    # except:
    #     pass

    today = (datetime.today()).replace(hour=0,minute=0,second=0,microsecond=0)
    holiday_file = (openpyxl.load_workbook(holiday_file_path)).active
    holiday_list = []
    next_birth_days= []
    for row in holiday_file.iter_rows(1, holiday_file.max_row):
        if isinstance(row[0].value,str):
            if 'holiday' not in row[0].value:
                row[0].value = datetime.strptime(row[0].value,'%d/%m/%Y')
        if isinstance(row[0].value,datetime):
            holiday_list.append(row[0].value)
    
    holiday_list.sort()
    next_day = today + timedelta(days=1)
    flag = False
    while not flag:
        if next_day in holiday_list:
            next_birth_days.append(next_day)
            next_day += timedelta(days=1)
        else:
            flag= True
    print(next_birth_days)

    for person in persons:
        birth_day = person['dob']
        message = None
        if isinstance(birth_day,str):
            birth_day = datetime.strptime(birth_day,'%d/%m/%Y')
        if birth_day.day == today.day and birth_day.month == today.month and person['status'] != 'done':
            print('hi')
            message = person['caption']
            image_path = "{}/{}".format(image_folder,person['image'])
        for day in next_birth_days:
            if birth_day.day == day.day and birth_day.month == day.month and person['status'] != 'done':
                message = 'advance ' + person['caption']
        if message:
            try:
                sendwhats_image(group_id,image_path,message,30,20)
                print('done')
                xml_file.cell(person['row'],max_column,'done')
            except:
                pass

    xml_book.save('details.xlsx')
    # start = "{}-{}-{} 09:00:00".format(next_day.year,next_day.month,next_day.day)
    # end = "{}-{}-{} 13:00:00".format(next_day.year,next_day.month,next_day.day)
    # print(start,end)
    # scheduler.add_job(func=task,args=[persons,image_folder,group_id],trigger='interval',hours=1,start_date=start,end_date=end,id='check')


@views.route('/schedule',methods=['GET'])
def schedule_task():
    if not scheduler.state:
        scheduler.start()

    with open('./file.txt','r') as text_file:
        text_line = ((text_file.readlines())[0]).split(':')
        file_path = text_line[len(text_line)-1]
    scope_app =['https://spreadsheets.google.com/feeds','https://www.googleapis.com/auth/drive'] 
    cred = ServiceAccountCredentials.from_json_keyfile_name('pywhatkit-sheet-f123bfb074c5.json',scope_app) 
    client = gspread.authorize(cred)
    file = client.open('pywhatkit')
    details_sheet = pd.DataFrame(file.get_worksheet(0).get_all_records())
    holiday_list=(pd.DataFrame(file.get_worksheet(1).get_all_records())).to_dict('list')
    settings =  (pd.DataFrame(file.get_worksheet(2).get_all_records())).to_dict('list')
    print(settings)
    holiday_list = holiday_list['holidays']
    for i in range(0,len(holiday_list)):
        holiday_list[i] =  datetime.strptime(holiday_list[i],'%d/%m/%y')
    persons = details_sheet.to_dict(orient='records')
    today = (datetime.today()).replace(hour=0,minute=0,second=0,microsecond=0)
    next_birth_days= []
    next_day = today + timedelta(days=1)
    flag = False
    while not flag:
        if next_day in holiday_list:
            next_birth_days.append(next_day)
            next_day += timedelta(days=1)
        else:
            flag= True
    print(next_birth_days)

    for person in persons:
        birth_day = person['date of birth']
        print(birth_day)
        message = None
        if isinstance(birth_day,str):
            try:
                birth_day = datetime.strptime(birth_day,'%d/%m/%Y')
            except:
                birth_day = datetime.strptime(birth_day,'%d/%m/%y')

        if birth_day.day == today.day and birth_day.month == today.month and person['status'] != 'done':
            print('hi')
            message = person['caption']
            image_path = "{}/{}".format(settings['images folder'][0],person['image'])
        for day in next_birth_days:
            if birth_day.day == day.day and birth_day.month == day.month and person['status'] != 'done':
                message = 'advance ' + person['caption']
        if message:
            try:
                # sendwhats_image(settings['group id'][0],image_path,message,30,20)
                print('done')
                person['status'] = 'don'
            except:
                pass
    # df.to_excel(file)
    # details_sheet.set_dataframes(df)
    df = pd.DataFrame.from_records(persons)
    change = file.get_worksheet(0)
    change.update([df.columns.values.tolist()]+df.values.tolist())
    return {'status':'success'}

# to start scheduler
@views.route('/schedule1',methods=['POST'])
def schedule_task1():
    # input data check
    data = request.json
    # start scheduler if not started
    if not scheduler.state:
        scheduler.start()

    if 'job' in data and data['job'] == 'stop':
        if scheduler.state:
            scheduler.shutdown()
            return {'status':'scheduled job is stopped successfully'}
        else:
            return {'status':'scheduled job is altready stopped'}

    if 'path' not in data or 'image folder' not in data or 'group id' not in data and 'holidays' not in data:
        return {'status':'check given details'}


    
    file_path = data['path']
    folder = data['image folder']
    group_id = data['group id']
    holiday_file_path = data['holidays']

    if 'job' in data and data['job'] == 'force start':
        on_time = datetime.now() + timedelta(minutes=1)
        print(on_time.hour,on_time.minute)
        scheduler.add_job(func=task,args=[folder,group_id,file_path,holiday_file_path], trigger='date',run_date=datetime(on_time.year, on_time.month, on_time.day, on_time.hour, on_time.minute,on_time.second),id='sudden dob message')
        return {'status':"force started your schedule today"}

    if not 'time' in data:
        return {'status':'please give time to schedule a job'}
    # schedule a job

    hour = data['time'][:2]
    minute = data['time'][3:5]
    scheduler.remove_all_jobs()
    scheduler.add_job(func=task,args=[folder,group_id,file_path,holiday_file_path], trigger='cron',day_of_week='mon-sun',hour=hour,minute=minute,id='dob message')
    return {'status':'your job was scheduled at {} everyday'.format(data['time'])}
